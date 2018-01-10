'''

This is what the script would look like if the payload is json. ISE 1.3 only supports XML so this is useful
only to test modifying user passwords on systems that support json.
Used for lab testing. Development done on ISE version 2.2

'''


import requests
import openpyxl
import json
import string
import secrets


startup_vars = check_startup()
server = startup_vars['server']
token = startup_vars['token']


workbook = 'user_file.xlsx'
sheet = 'Sheet1'


call = '/ers/config/internaluser'
user_filter = '?filter=name.EQ.{}'


headers = {
    "authorization": "Basic " + token,
    "accept": "application/json",
    "content-type": "application/json"
}



headers = {"accept": "application/json","content-type": "application/json"}

def random_pwd() :


    alphabet = string.ascii_letters + string.digits + string.punctuation
    alphabet = alphabet.replace('^', '')
    alphabet = alphabet.replace('`', '')
    alphabet = alphabet.replace('"', '')
    alphabet = alphabet.replace("'", '')
    print(alphabet)


    while True:
        passwd_str = ''.join(secrets.choice(alphabet) for i in range(10))
        print('sum is: ', sum(c.isdigit() for c in passwd_str))
        print('password is: ', passwd_str)
        if (any(c.islower() for c in passwd_str)
            and any(c.isupper() for c in passwd_str)
            and any(c in string.punctuation for c in passwd_str)
            and sum(c.isdigit() for c in passwd_str) >= 3
            and sum(c in string.punctuation for c in passwd_str) <= 3
            and passwd_str[0] in (string.ascii_letters + string.digits)):
            print(passwd_str)
            break

    return passwd_str

def update_user(user_id, name) :

    new_passwd = random_pwd()

    req_body_json = """  {{
        "InternalUser" : {{
            "id" : "{}",
            "description" : "{}",
            "name" : "{}",
            "password" : "{}",
            "enablePassword" : "{}",
            "customAttributes" : {{
            }}
        }}
    }}
    """.format(user_id, new_passwd, name, new_passwd, new_passwd)

    url = 'https://' + server + ':9060' + call + "/{}".format(user_id)

    print('New password is:', new_passwd)
    print('req_body_json is:', req_body_json)
    print('headers from main: ', headers)
    print('url is: ', url)

    resp = requests.put(url, data=req_body_json, headers=headers, verify=False)
    response_json = resp.json()  # Get the json-encoded content from response
    print("Status.put: ", resp.status_code)  # This is the http request status
    print("Status.put.Headers: ", resp.headers)
    print(json.dumps(response_json,
                     indent=4))  # Convert "response_json" object to a JSON formatted string and print it out


    return new_passwd

wb = openpyxl.load_workbook(workbook)
sheet = wb.get_sheet_by_name(sheet)
data_list = []
for row in range(2, sheet.max_row + 1):
    data = {}
    first = sheet['A' + str(row)].value
    last = sheet['B' + str(row)].value
    name = sheet['C' + str(row)].value
    mail = sheet['D' + str(row)].value
    data['first'] = first
    data['last'] = last
    data['name'] = name
    data['mail'] = mail
    print(data)
    data_list.append(data)

print(data_list)

users=[]
for index in range(len(data_list)) :
    print(data_list[index]['mail'])
    print(data_list[index]['name'])
    users.append(data_list[index]['name'])

print(users)


for user in range(len(users)) :
    print('user is: ', users[user])
    url = 'https://' + server + ':9060' + call + user_filter.format(users[user])
    print(url)
    print('Headers is: ', headers)
    print('url is: ', url)

    resp = requests.get(url, headers=headers, verify=False)
    response_json = resp.json()  # Get the json-encoded content from response
    print("Status: ", resp.status_code)  # This is the http request status
    print("Status.Headers: ", resp.headers)
    print(json.dumps(response_json,
                     indent=4))  # Convert "response_json" object to a JSON formatted string and print it out

    total = response_json['SearchResult']['total']


    if total == 1 :
        print(users[user], 'exists in ISE')
        id = response_json['SearchResult']['resources'][0]['id']
        print('id is:', id, 'and user is:', users[user])
        new_passwd = update_user(id, users[user])
        print('from main:', new_passwd)
        print('name for row', user, ' is', users[user])
        print('row is', user)
        sheet.cell(row=user+2, column=5).value = new_passwd

    elif total == 0 :
        print(users[user], 'not defined in ISE')

wb.save('user_file.xlsx')

