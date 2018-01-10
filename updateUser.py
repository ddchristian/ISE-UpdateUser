'''

This is the main script written to iterate XML data.
Supported with ISE 1.3 and above where XML is the only option
Initially developed and tested against ISE version 2.2 server

'''

import os
import requests
import openpyxl
import string
import secrets
import xml.etree.ElementTree as ET
from openpyxl.utils import get_column_letter, column_index_from_string
from xml.dom import minidom
from startup import check_startup


startup_vars = check_startup()
server = startup_vars['server']
token = startup_vars['token']


URL = 'https://' + server + ':9060'
base_call = '/ers/config/internaluser'


headers = {}
accept = 'application/vnd.com.cisco.ise.identity.internaluser.1.1+xml;charset=UTF-8'
headers['authorization'] = 'Basic {}'.format(token)
headers['accept'] = accept

all_users = []


def get_user_id():

    url = URL + base_call

    print ('url is: ', url)
    resp = requests.get(url, headers=headers, verify=False)
    print('resp.headers[link] is: \n', resp.headers['Content-Length'])
    print(resp.content)

    root = ET.fromstring(resp.content)
    xmlstr = minidom.parseString(ET.tostring(root)).toprettyxml(indent="   ")
    #print('xml string is: \n', xmlstr)

    user_id = []
    for child in root.iter(tag='{ers.ise.cisco.com}resource'):
        #for key, value in child.attrib.items():
        #    print(key, ':', value)
        #print(child.attrib['id'])
        user_id.append(child.attrib['id'])

    print(user_id)

    return user_id


def get_user_detail(user_id):

    user_detail = {}

    url = URL + base_call + '/' + user_id
    print(url)

    resp = requests.get(url, headers=headers, verify=False)
    print('resp.headers[link] is: \n', resp.headers['Content-Length'])
    print(resp.content)

    root = ET.fromstring(resp.content)
    xmlstr = minidom.parseString(ET.tostring(root)).toprettyxml(indent="   ")
    #print('xml string is: \n', xmlstr)

    print('user name=', root.attrib['name'])
    user_detail['name'] = root.attrib['name']
    print('user desctiption=', root.attrib['description'])
    user_detail['description'] = root.attrib['description']
    print('user id=', root.attrib['id'])
    user_detail['id'] = root.attrib['id']

    for child in root.iter('*'):
        if child.tag in ['firstName', 'lastName', 'email']:
            #print('Tag=', child.tag, 'Value=', child.text)
            user_detail[child.tag] = child.text

    #print(user_detail)

    return user_detail


def random_pwd() :

    alphabet = string.ascii_letters + string.digits + string.punctuation
    alphabet = alphabet.replace('^', '')
    alphabet = alphabet.replace('`', '')
    alphabet = alphabet.replace('"', '')
    alphabet = alphabet.replace("'", '')
    alphabet = alphabet.replace(",", '')
    alphabet = alphabet.replace(".", '')
    alphabet = alphabet.replace("\\", '')
    alphabet = alphabet.replace("~", '')
    alphabet = alphabet.replace("I", '')
    alphabet = alphabet.replace("l", '')
    alphabet = alphabet.replace("O", '')
    print(alphabet)

    while True:
        passwd_str = ''.join(secrets.choice(alphabet) for i in range(10))
        #print('sum is: ', sum(c.isdigit() for c in passwd_str))
        #print('password is: ', passwd_str)
        if (any(c.islower() for c in passwd_str)
            and any(c.isupper() for c in passwd_str)
            and any(c in string.punctuation for c in passwd_str)
            and sum(c.isdigit() for c in passwd_str) >= 2
            and sum(c in string.punctuation for c in passwd_str) <= 2
            and passwd_str[0] in (string.ascii_letters + string.digits)):
            break

        #print(passwd_str)

    return passwd_str


def update_user(user_name, workbook):

    contentType = 'application/vnd.com.cisco.ise.identity.internaluser.1.1+xml;charset=UTF-8'
    headers['content-type'] = '{}'.format(contentType)
    #print('headers is: ', headers)

    user_filter = '?filter=name.EQ.{}'
    url_user_filter = URL + base_call + user_filter.format(user_name)
    #url = URL + base_call + '/' + user_name
    #print(url_user_filter)

    #workbook ='user_file.xlsx'
    sheet = 'Sheet1'

    print(url_user_filter)
    print('Headers is: ', headers)
    print('url is: ', url_user_filter)

    wb = openpyxl.load_workbook(workbook)
    sheet = wb.get_sheet_by_name(sheet)
    data_list = []
    for row in range(2, sheet.max_row + 1):
        data = {}
        first = sheet['A' + str(row)].value
        last = sheet['B' + str(row)].value
        name = sheet['C' + str(row)].value
        mail = sheet['D' + str(row)].value
        data['first'] = first
        data['last'] = last
        data['name'] = name
        data['mail'] = mail
        print(data)
        data_list.append(data)

    print(data_list)

    resp = requests.get(url_user_filter, headers=headers, verify=False)
    print('resp.headers[link] is: \n', resp.headers['Content-Length'])
    print(resp.content)

# check if request returned a valid user. Else exit with error statement

    root = ET.fromstring(resp.content)
    xmlstr = minidom.parseString(ET.tostring(root)).toprettyxml(indent="   ")
    #print('xml string is: \n', xmlstr)

    #print('root=', root)

    if int(root.attrib['total']) > 0:
        for child in root.iter(tag='{ers.ise.cisco.com}resource'):
            user_id = child.attrib['id']
            print('user_id=', user_id)

        url_user_detail = URL + base_call + '/' + user_id

        print('Headers is: ', headers)
        print('url is: ', url_user_detail)

        resp = requests.get(url_user_detail, headers=headers, verify=False)
        print('resp.headers[link] is: \n', resp.headers['Content-Length'])
        print(resp.content)

        root = ET.fromstring(resp.content)
        xmlstr = minidom.parseString(ET.tostring(root)).toprettyxml(indent="   ")

        new_password = random_pwd()

        root.attrib['description'] = new_password
        for child in root.iter('*'):
            if child.tag in ['enablePassword', 'password']:
                child.text = new_password

        xmlstr = minidom.parseString(ET.tostring(root)).toprettyxml(indent="   ")
        #print('xml string is: \n', xmlstr)

        resp = requests.put(url_user_detail, xmlstr, headers=headers, verify=False)

        print('resp.status is: \n', resp.status_code)

        root = ET.fromstring(resp.content)
        xmlstr = minidom.parseString(ET.tostring(root)).toprettyxml(indent="   ")
        print('xml string is: \n', xmlstr)

    else:
        new_password = 'NOT_FOUND_a1'

    return new_password


def select_get_ise_users(workbook):

    if workbook == '':
        workbook = 'import_user_file.xlsx'

    if not os.path.isfile(workbook):

        if not '.' in workbook :
            print('Renamed ', workbook, 'to ', workbook, '\b.xlsx')
            workbook = workbook + '.xlsx'

        if not workbook.split('.')[1] == 'xlsx':
            workbook = workbook.split('.')[0] + '.xlsx'
            print('Invalid file extensions. Changing to . xlsx for file {}.'.format(workbook))

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'Sheet1'

        sheet['A1'].value = 'First'
        sheet['B1'].value = 'Last'
        sheet['C1'].value = 'UserId'
        sheet['D1'].value = 'email'
        sheet['E1'].value = 'passwd'
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 10
        sheet.column_dimensions['D'].width = 18
        sheet.column_dimensions['E'].width = 12
        sheet.sheet_view.zoomScale = 150

        # print('workbook =', workbook)
        # print('sheet =', sheet)
    else:
        sheet = 'Sheet1'
        wb = openpyxl.load_workbook(workbook)
        sheet = wb.get_sheet_by_name(sheet)


    for user in get_user_id():
        all_users.append(get_user_detail(user))

    #print(all_users)

    print('Total users found in ISE database =', len(all_users))

    print('Start writing to excel')
    #workbook = 'import_user_file.xlsx'
    count = 2
    for user in all_users:
        sheet['A' + str(count)].value = user['firstName']
        sheet['B' + str(count)].value = user['lastName']
        sheet['C' + str(count)].value = user['name']
        sheet['D' + str(count)].value = user['email']
        count = count + 1
    wb.save(workbook)

    return

# After excel list has been check and edited now read it back again to make changes

def select_update_password(workbook):

    print('Start reading from excel')
    #workbook = 'user_file.xlsx'

    if workbook == '':
        workbook = 'user_file.xlsx'

    if workbook.find('.') == -1:
        print('Renamed ', workbook, 'to ', workbook, '\b.xlsx')
        workbook = workbook + '.xlsx'

    if (not os.path.isfile(workbook)) and (not workbook.split('.')[1] == 'xlsx'):
        print('The file {} does not exist or is invalid. '
              'Check the file name and try again.\nExiting application....'.format(workbook))
        exit()
    sheet = 'Sheet1'

    wb = openpyxl.load_workbook(workbook)
    sheet = wb.get_sheet_by_name(sheet)

    names = {}
    for row in range(2, sheet.max_row + 1):
        name = sheet['C' + str(row)].value
        new_password = update_user(name, workbook)
        #print('name=', name, 'new password=', new_password)
        sheet.cell(row=row, column=column_index_from_string('E')).value = new_password
        names[name]=new_password

    wb.save(workbook)
    print(names)

    return


print('A. Get user database from ISE and write to excel sheet.')
print('B. Update user passwords.')
selection = input('Input:')

if selection.strip().lower() == 'a':
    print('Selection = A')
    workbook = input('Select the file to write to in format file_name.xlsx.\n'
                     'Hit enter to use default import_user_file.xlsx:')
    select_get_ise_users(workbook)
elif selection.strip().lower() == 'b' :
    print('Selection = B')
    workbook = input('Select the file to write to in format file_name.xlsx.\n'
                     'Hit enter to use default user_file.xlsx:')
    select_update_password(workbook)


